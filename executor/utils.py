import os
import xlsxwriter
from openpyxl import load_workbook


def append_row_to_xlsx(file_name, row_data):
    # 检查文件是否存在
    if os.path.exists(file_name):
        # 如果文件存在，使用 openpyxl 读取内容
        wb = load_workbook(file_name)
        ws = wb.active
        # 找到当前的最大行数
        max_row = ws.max_row
        # 将新数据写入到下一行
        for col_num, data in enumerate(row_data, start=1):
            ws.cell(row=max_row + 1, column=col_num, value=data)
        # 保存更改
        wb.save(file_name)
    else:
        # 如果文件不存在，使用 xlsxwriter 创建新文件
        with xlsxwriter.Workbook(file_name) as workbook:
            worksheet = workbook.add_worksheet()
            # 写入数据
            for col_num, data in enumerate(row_data, start=0):
                worksheet.write(0, col_num, data)  # 在第一行写入数据


if __name__ == '__main__':
    data = {
        "name": "alex",
        "age": 18,
        "phone": "15171618793"
    }
    data_list = list(data.values())
    append_row_to_xlsx("./test.xlsx", data_list)
